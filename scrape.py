import pandas as pd
import requests as req
import hashlib
import datetime
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import numpy as np

class ParseFile:

    def __init__(self, file):
        self.header_dates = []
        self.header_url = []
        self.hash_values = []
        self.hash_url = []
        self.last_modified_url = []
        self.current_date = []
        self.file = file
        self.hash_df = pd.DataFrame
        self.header_df = pd.DataFrame
        self.merge_df = pd.DataFrame
        self.df = pd.DataFrame
        self.hash_frame = pd.DataFrame
        self.check_hash_against = []
        self.check_hash_url = []
        self.check_hash_date =[]
        self.checked_hash_df = pd.DataFrame
        self.current_df = pd.DataFrame
        self.hold_list = []
        self.check_hash_list = []
        self.master_hash_list = []
        self.final_list = []
        self.merge_final = pd.DataFrame
        self.olddate_list = []
        self.check_hash_time = []
        self.master_hash_time = []
        self.holdme = []

    def set_file(self, file):
        """Sets file to be parsed"""
        self.file = file
        return self.set_file

    def set_df(self):
        self.current_df = pd.read_excel(self.file)


    def initial_parse(self):
        count = 0
        self.df = pd.read_excel(self.file)
        for url in self.df['Website URL']:
            r = req.get(url, verify = False)
            h = hashlib.sha512(r.text.encode('utf-8)'))
            h = h.hexdigest()
            count +=1
            self.hash_values.append(h)
            self.hash_url.append(url)
            self.current_date.append(datetime.datetime.now())

        # print(count, self.hash_values)
        # print(count, self.hash_url)
        # print(count, self.current_date)
        return self.header_dates, self.header_url, self.hash_values, self.hash_url, self.current_date

    def final_check(self, file1, file2):
        """Need to fix last mod changign even when hash values are the same.
                Also different hash values occuring, but last mod becoming 'nan' as opposed to current date. """
        self.file= file1
        self.file2 = file2
        self.frame_1 = pd.read_excel(self.file)  # Read data into pandas DF
        for i in self.file['Hash-Value']:
            if i in self.file2['Hash-Value']:
                self.final_list.append(i) #+ matching hash values from "MasterFile" and "checkAgainst"
            else:
                self.olddate_list.append(datetime.datetime.now())
        self.checked_hash_df = pd.DataFrame({"Last-Modified": self.olddate_list, "Hash-Value": self.final_list})  # Matching hash values from both files
        return self.checked_hash_df














# , "Last-Modified": self.current_date 'Website URL': self.hash_url,
    def create_hash_df(self):
        self.hash_df = pd.DataFrame({'Website URL': self.hash_url, "Hash-Value": self.hash_values, 'Last-Modified': self.current_date})
        return self.hash_df

    def combine_df(self): # lexicographically
        self.merge_df = self.df.astype(str).merge(self.hash_df.astype(str), on=['Hash-Value'],how='right', suffixes=('_', ''))
        # self.merge_df = self.merge_df.drop(columns=["Last-Modified_"])
        return self.merge_df





    # def concat_df(self): # lexicographically
    #     self.res = self.hash_df.astype(str).set_index(['Website URL', 'Hash-Value']) \
    #         .combine_first(self.df.astype(str).set_index(['Website URL', 'Hash-Value'])) \
    #         .reset_index()
    #     return self.res

    def final_merge(self):
        self.merge_final = self.merge_df.astype(str).merge(self.df.astype(str), on=['Website URL', 'Hash-Value'], how='left', suffixes=('_', ''))
        # self.merge_final = self.merge_final.drop(columns=["Last-Modified_"])
        self.merge_final = self.merge_final.astype(str)
        return self.merge_final

    def save_wb(self, save_file, use_df):
            self.save_to_file = save_file
            self.use_df = use_df
            wb = Workbook()
            ws = wb.active

            for r in dataframe_to_rows(self.use_df, index=True, header=True ):
                ws.append(r)

            for cell in ws['A'] + ws[1]: # Bolds header and index
                cell.style = 'Pandas'
            ws.delete_rows(2, 1) # empty row that appends
            wb.save(self.save_to_file)

    def check_hash(self):
        """Need to fix last mod changign even when hash values are the same.
        Also different hash values occuring, but last mod becoming 'nan' as opposed to current date. """
        self.hash_frame = pd.read_excel(self.file) # Read data into pandas DF
        for hash_val in self.hash_frame['Hash-Value']:  # Obtain all hashes in Hash-Value column
            if not hash_val != hash_val: # Where hash-value is given (any valid value)
                if self.file == "checkAgainst.xlsx":
                    self.check_hash_list.append(hash_val) # append all hash values
                elif self.file == "CheckV1.xlsx":
                    self.master_hash_list.append(hash_val)
        return self.check_hash_list, self.master_hash_list

    def compare_list(self):

        for i in self.check_hash_list:
            if i in self.master_hash_list:
                self.final_list.append(i) #+ matching hash values from "MasterFile" and "checkAgainst"
            else:
                self.olddate_list.append(datetime.datetime.now())
        self.checked_hash_df = pd.DataFrame({"Last-Modified": self.olddate_list, "Hash-Value": self.final_list})  # Matching hash values from both files
        return self.checked_hash_df



def main():
    file = ParseFile("MasterFileTesting.xlsx")  # instantiate class instance
    file.set_df() # Sets file as current DF
    file.initial_parse()
    file.create_hash_df() # wrongfully overwriting date for same hash
    file.save_wb("checkHashDF.xlsx", file.hash_df)
    file.combine_df() # Used to obtain combine_df2
    file.save_wb("merge_hashMaster.xlsx", file.merge_df) # should not overwrite previous date if hash unchanged
    # file.concat_df()
    # file.save_wb("merge_hashMaster2.xlsx", file.res)
    # file.combine_df2() # Issue in merge is here
    # file.save_wb("checkOriginal.xlsx", file.df) # correct original output. unnecssary
    # file.save_wb("checkAgainst.xlsx", file.merge_df2) # last mod date not staying the same even when hash values are same
    # file.set_file("checkAgainst.xlsx")
    # file.check_hash() # Checks current hash value from MasterFile
    # file.set_file("CheckV1.xlsx")
    # file.check_hash()
    # file.compare_list() #initial run will be empty
    # file.final_check("merge_hashMaster.xlsx", "checkHashDF.xlsx")
    # file.final_merge()
    # file.save_wb('FinalDF.xlsx', file.merge_final)


"""Taking the time from check3 and making it the official end time. not wanted"""
"""Add the last modified time for new appended hash values into officiali list and done."""
main()
