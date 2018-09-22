import pandas as pd
import requests as req
import numpy as np
import openpyxl

# Read an Excel file into a pandas DataFrame
df = pd.read_excel('cryptofundurls.xlsx', sheet_name='Sheet1')
df['Last-Modified'] = np.nan  # Sets all rows as empty


for value in df['Website URL']:  # Obtain all values in website column
    r = req.get(value) # Response object
    if 'Last-Modified' in r.headers: # Checks headers for string
        contents = r.headers['Last-Modified'] # Sets 'contents' as modified date
        print(contents)
        # print(type(contents)) # Type string
        # print(type(r.headers)) # requests.structures.CaseInsensitiveDict
d = []
for p in game.players.passing():
    d.append({'Player': p, 'Team': p.team, 'Passer Rating':
        p.passer_rating()})

pd.DataFrame(d)







# To convert a dataframe into a worksheet highlighting the header and index:
# wb = Workbook()
# ws = wb.active
#
# for r in dataframe_to_rows(df, index=True, header=True):
#     ws.append(r)
#
# for cell in ws['A'] + ws[1]:
#     cell.style = 'Pandas'
#
# wb.save("pandas_openpyxl.xlsx")
