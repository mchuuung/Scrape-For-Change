import pandas as pd
import requests as req
import hashlib
import datetime
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows


class ParseFile:

    def __init__(self, file):
        self.header_dates = []
        self.header_url = []
        self.hash_values = []
        self.hash_url = []
        self.last_modified_url = []
        self.current_date = []
        self.file = file
        self.hash_df = pd.DataFrame
        self.header_df = pd.DataFrame
        self.merge_df = pd.DataFrame
        self.df = pd.DataFrame
        self.hash_frame = pd.DataFrame
        self.check_hash_against = []
        self.check_hash_url = []
        self.check_hash_date =[]
        self.checked_hash_df = pd.DataFrame
        self.current_df = pd.DataFrame
        self.hold_list = []
        self.check_hash_list = []
        self.master_hash_list = []
        self.final_list = []
        self.merge_final = pd.DataFrame
        self.olddate_list = []
        self.check_hash_time = []
        self.master_hash_time = []
        self.holdme = []

    def set_file(self, file):
        """Sets file to be parsed"""
        self.file = file
        return self.set_file

    def set_df(self):
        self.current_df = pd.read_excel(self.file)


    def get_file(self):
        count = 0
        count_not_same = 0
        """Parses initial file. Sends http request to determine if 'Last-Modified' is given.
        Appends 'Last-Modified' date if given, else sets current date and hash of website."""
        self.df = pd.read_excel(self.file) # Read an Excel file into a pandas DataFrame
        for url in self.df['Website URL']:  # Obtain all url's in website column
            r = req.get(url, verify=False)  # Response object. Value is set to url of website
            if 'Last-Modified' in r.headers:  # Checks headers for string
                header_date = r.headers['Last-Modified']  # Sets 'header_date' as modified date
                website_url = url  # url's where website has been last modified
                self.header_dates.append(header_date)  # list of modified dates
                self.header_url.append(website_url)  # list of urls
            else:
                h = hashlib.sha512(r.text.encode('utf-8')) # hash the web page
                h = h.hexdigest()
                self.hash_values.append(h)
                self.hash_url.append(url)

        for y in self.df['Hash-Value']:
            if y in self.hash_values:
                count+=1

            else:
                count_not_same+=1
                # self.current_date.append(datetime.datetime.now())
        print(count)
        print(count_not_same)



        # print(count, self.current_date)
        return self.header_dates, self.header_url, self.hash_values, self.hash_url, self.current_date

    def create_hash_df(self):
        self.hash_df = pd.DataFrame({'Website URL': self.hash_url,  "Hash-Value": self.hash_values})  # date of last modified and time
        return self.hash_df

    def create_header_df(self):
        self.header_df = pd.DataFrame({'Website URL': self.header_url, 'Last-Modified': self.header_dates})  # date of last modified and time
        return self.header_df

    def combine_df(self):
        self.merge_df = self.hash_df.astype(str).merge(self.header_df.astype(str), on=['Last-Modified', 'Website URL'], how='outer', suffixes=('_', ''))
        return self.merge_df

    def combine_df2(self):
        self.merge_df2 = self.df.astype(str).merge(self.merge_df.astype(str), on=['Website URL'],how='left', suffixes=('_', ''))
        self.merge_df2 = self.merge_df2.drop(columns=["Hash-Value_"])
        #         self.merge_df2 = self.df.astype(str).merge(self.merge_df.astype(str), on=['Last-Modified', 'Website URL', 'Hash-Value'],how='left', suffixes=('_', '')) Returns copy of original
        #         self.merge_df2 = self.df.astype(str).merge(self.merge_df.astype(str), on=['Last-Modified', 'Website URL', 'Hash-Value'],how='right', suffixes=('_', '')) Values of new merge without hjeader
        return self.merge_df2

    # def create_original(self):
    #     self.df = pd.read_excel(self.file)

    def final_merge(self):
        self.merge_final = self.merge_df2.astype(str).merge(self.checked_hash_df.astype(str), on=['Hash-Value', 'Last-Modified'], how='left', suffixes=('_', ''))
        # self.merge_final = self.merge_final.drop(columns=["Last-Modified_"])
        self.merge_final = self.merge_final.astype(str)
        return self.merge_final

    def save_wb(self, save_file, use_df):
            self.save_to_file = save_file
            self.use_df = use_df
            wb = Workbook()
            ws = wb.active

            for r in dataframe_to_rows(self.use_df, index=True, header=True ):
                ws.append(r)

            for cell in ws['A'] + ws[1]: # Bolds header and index
                cell.style = 'Pandas'
            ws.delete_rows(2, 1) # empty row that appends
            wb.save(self.save_to_file)

    def check_hash(self):
        """Need to fix last mod changign even when hash values are the same.
        Also different hash values occuring, but last mod becoming 'nan' as opposed to current date. """
        self.hash_frame = pd.read_excel(self.file) # Read data into pandas DF
        for hash_val in self.hash_frame['Hash-Value']:  # Obtain all hashes in Hash-Value column
            if not hash_val != hash_val: # Where hash-value is given (any valid value)
                if self.file == "checkAgainst.xlsx":
                    self.check_hash_list.append(hash_val) # append all hash values
                elif self.file == "CheckV1.xlsx":
                    self.master_hash_list.append(hash_val)
        return self.check_hash_list, self.master_hash_list

    def compare_list(self):

        for i in self.check_hash_list:
            if i in self.master_hash_list:
                self.final_list.append(i) #+ matching hash values from "MasterFile" and "checkAgainst"
            else:
                self.olddate_list.append(datetime.datetime.now())
        self.checked_hash_df = pd.DataFrame({"Last-Modified": self.olddate_list, "Hash-Value": self.final_list})  # Matching hash values from both files
        return self.checked_hash_df



def main():
    file = ParseFile("MasterFileTesting.xlsx")  # instantiate class instance
    file.set_df() # Sets file as current DF
    file.save_wb('CheckV1.xlsx', file.current_df) # Store current master file. unnecessary
    file.set_file("MasterFileTesting.xlsx")
    file.get_file()
    # file.create_hash_df() # wrongfully overwriting date for same hash
    # file.save_wb("checkHashDF.xlsx", file.hash_df)
    # file.create_header_df() # correct values remaining
    # file.save_wb("checkHeaderDF.xlsx", file.header_df)
    # file.combine_df() # Used to obtain combine_df2
    # file.save_wb("checkHash_HeaderDF.xlsx", file.merge_df) # should not overwrite previous date if hash unchanged
    # file.combine_df2() # Issue in merge is here
    # file.save_wb("checkOriginal.xlsx", file.df) # correct original output. unnecssary
    # file.save_wb("checkAgainst.xlsx", file.merge_df2) # last mod date not staying the same even when hash values are same
    # file.set_file("checkAgainst.xlsx")
    # file.check_hash() # Checks current hash value from MasterFile
    # file.set_file("CheckV1.xlsx")
    # file.check_hash()
    # file.compare_list() #initial run will be empty
    # file.final_merge()
    # file.save_wb('MasterFileTesting.xlsx', file.merge_final)


"""Taking the time from check3 and making it the official end time. not wanted"""
"""Add the last modified time for new appended hash values into officiali list and done."""
main()